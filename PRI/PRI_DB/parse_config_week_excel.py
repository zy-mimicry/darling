# coding = 'utf-8'
from __future__ import unicode_literals
import xlrd
import os
import sys
import argparse
import django
from openpyxl import load_workbook
 
pathname = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, pathname)
sys.path.insert(0, os.path.abspath(os.path.join(pathname, '..')))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "PRI_System.settings")
django.setup()

from PRI_DB.models import *  
reload(sys)  
sys.setdefaultencoding('utf-8')
import jira
sys.path.insert(0, os.path.abspath(os.path.join(pathname, '../../lib')))
from myjira import *
import time, datetime
        
def read_excel(file):
    try:
        data = load_workbook(file)
        return data
    except Exception as err:
        print(err)
        
def covert_hypelink_format(hypelink):
    #'=HYPERLINK("https://skutracker.sierrawireless.local/projects/2098", "2098")'
    #=A4("https://agile.sierrawireless.com/Agile/PLMServlet?action=OpenEmailObject&classid=8000&objid=11782795", "TDN-011787")
    if 'HYPERLINK' in hypelink:
        temp = hypelink.split('HYPERLINK')[1].split(',')
        link = temp[0].strip('(')
        value = temp[1].strip(')')
    elif 'hyperlink' in hypelink:
        temp = hypelink.split('hyperlink')[1].split(',')
        link = temp[0].strip('(')
        value = temp[1].strip(')')        
    elif 'A4' in hypelink:
        temp = hypelink.split('A4')[1].split(',')
        link = temp[0].strip('(')
        value = temp[1].strip(')')
    else:
        link = 'NULL'
        hypelink.strip()
        #agile maybe is null and has no hypelink
        return link, hypelink
        
    value = value.replace("\"", "").strip()
    link = link.replace("\"", "")
    return link, value
    
def covert_time_format(value):
    if isinstance(value, unicode) == True and value != 'None':
        value = str(value).split(' ')[0]
        date = str(value).strip().replace('/', '-')
        if '-' not in date:#no date but has str
            return 'None'
        #2017-9-18-ok
        temp = date.split('-')
        data = temp[0]+'-'+temp[1]+'-'+temp[2]
        return data
        
    #value: 2017/11/10 fw download fail, handup
    try:
        data = xlrd.xldate_as_tuple(value, 0)
        date = str(data[0])+'-'+str(data[1])+'-'+str(data[2])
    except:
        date = str(value).strip().replace('/', '-')
    
    if date != 'None':
        try:
            dt = datetime.datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
            dt = dt.strftime("%Y-%m-%d")
        except:
            return date 
    else:
        return date 
    
    return dt  
    
def excel_to_dict(read_excel_file, sheet_name):
    data = read_excel(read_excel_file)
    table = data[sheet_name]
    rows = table.max_row
    data_lists = []
    
    if sheet_name == 'OEMPRI-Status' or sheet_name == 'FWTOOLS-Status':
        cols = 11
    elif sheet_name == 'Factory-Status':
        cols = 15
    elif sheet_name == 'Flow-Status':
        cols = 18

    for i in range(3, rows+1):
        data_dict = {}
        for j in range(1, cols):
            title = table.cell(2, j).value
            value = table.cell(i, j).value
            data_dict[title] = value
        data_lists.append(data_dict)

    return data_lists
    
def base_info_to_mysql(dicts):
    for dict in dicts:
        item_dict = {}
        oempri_link,oempri_value = covert_hypelink_format(dict['Key'])
        list=Jira_List.objects.filter(Key=oempri_value).all()
        
        if len(list) == 0:
            item_dict['Key'] = oempri_value
            item_dict['Summary'] = dict['Summary']
            item_dict['Status'] = dict['Status']
            item_dict['CreatedDate'] = covert_time_format(dict['Created'])
            item_dict['DueDate'] = covert_time_format(dict['Due Date'])
            item_dict['Components'] = dict['Component/s']
            item_dict['Assignee'] = dict['Assignee']
            item_dict['Reporter'] = dict['Reporter']
            item_dict['Week'] = covert_time_format(dict['Week'])
            item_dict['PlanDueDate'] = covert_time_format(dict['Plan due date'])
            print 'base_info_to_mysql:%s' %oempri_value
            Jira_List.objects.create(**item_dict)
        else:
            print 'the data is repeated: %s'%oempri_value
 
def factory_info_to_mysql(dicts):
    for dict in dicts:
        item_dict = {}
        oempri_link,oempri_value = covert_hypelink_format(dict['OEMPRI'])
        sku_link,sku_value = covert_hypelink_format(dict['SKUTracker'])
        agile_link,agile_value = covert_hypelink_format(dict['Agile'])
        list=Jira_List.objects.filter(Key=oempri_value).all()

        if dict['Factory'] == True:
            item_dict['IsFactory'] = 'TRUE'
        item_dict['SkuNumber'] = str(dict['SKU Number']).split('.')[0]
        item_dict['PartNumber'] = str(dict['Part Number']).split('.')[0]
        item_dict['ProjID'] = sku_value
        item_dict['AgileID'] = agile_value
        item_dict['AgileUrl'] = agile_link
        item_dict['Customers'] = dict['Customers']
        item_dict['WorkPackage'] = dict['Work Package']
        if len(list) == 0:
            item_dict['Key'] = oempri_value
            item_dict['Status'] = dict['Status']
            item_dict['CreatedDate'] = covert_time_format(dict['Created'])
            item_dict['DueDate'] = covert_time_format(dict['Due Date'])
            item_dict['Components'] = dict['Component/s']
            item_dict['Assignee'] = dict['Assignee']               
            item_dict['Reporter'] = dict['Reporter']
            print 'in factory status but not in oempri status: %s' %oempri_value, oempri_link
            Jira_List.objects.create(**item_dict)
        else:
            Jira_List.objects.filter(Key=oempri_value).update(**item_dict)
            
def flow_info_to_mysql(dicts):
    for dict in dicts:
        item_dict = {}
        oempri_link,oempri_value = covert_hypelink_format(dict['Key'])
        list=Jira_List.objects.filter(Key=oempri_value).all()
        
        item_dict['issueType'] = dict['Issue Type']
        item_dict['Open'] = dict['OPEN']
        item_dict['InProcess'] = dict['IN PROCESS']
        item_dict['Generated'] = dict['GENERATED']               
        item_dict['Tested'] = dict['TESTED']
        item_dict['Reviewed'] = dict['REVIEWED']
        item_dict['Integrated'] = dict['INTERGRATED']
        item_dict['Validated'] = dict['VALIDATED']
        item_dict['AllFilesAdded'] = dict['ALL FILES ADDED']
        item_dict['SpkgValidated'] = dict['SPKG VALIDATED']
        item_dict['LogValidated'] = dict['LOG VALIDATED']
        item_dict['Closed'] = dict['CLOSED']
            
        if len(list) == 0:
            item_dict['Key'] = oempri_value
            item_dict['Status'] = dict['STATUS']
            item_dict['Assignee'] = dict['Assignee']
            item_dict['Reporter'] = dict['Reporter']
            item_dict['CreatedDate'] = covert_time_format(dict['Created'])
            print 'in flow status but not in oempri status: %s' %item_dict['Key']
            Jira_List.objects.create(**item_dict)
        else:
            Jira_List.objects.filter(Key=oempri_value).update(**item_dict)

def qti_to_mysql(filter, table_name):
    print 'start process qti_to_mysql...'
    prits=find_jira_oempri(jql_str=filter, cmptool=False)
    
    for id in prits:
        info = prits[id]
        list=table_name.objects.filter(Key=id).all()
        item_dict = {}
        temp_str = ''
        item_dict['Key'] = info['OempriId']
        item_dict['Summary'] = info['Summary']
        for version in info['FixVersion']:
            temp_str += str(version) + ','
        item_dict['FixVersion'] = temp_str[:-1]
        item_dict['Status'] = info['Status']
        item_dict['CreatedDate'] = info['Created'].replace('/', '-')
        if info['Due Date'] != None:
            item_dict['DueDate'] = info['Due Date'].replace('/', '-')
        else:
            item_dict['DueDate'] = info['Due Date']
        item_dict['Assignee'] = info['Assignee']
        item_dict['Reporter'] = info['Reporter']
        if info['Updated'] != None:
            item_dict['UpdateDate'] = info['Updated'].replace('/', '-')
        else:
            item_dict['UpdateDate'] = info['Updated']
        item_dict['Components'] = info['Components']
        #print item_dict
        
        if len(list) == 0:
            table_name.objects.create(**item_dict)
        else:
            table_name.objects.filter(Key=info['OempriId']).update(**item_dict)
    print 'exit process qti_to_mysql.'
                    
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process input arguments for modify.py")
    parser.add_argument('-p', '--Path', type=str, help='the configuration-weekly file path')
    args = parser.parse_args()
    excel_path = args.Path #such as D:\PRI_System\PRI_DB\configuration-weekly.xlsx
    
    dicts = excel_to_dict(excel_path, 'OEMPRI-Status')
    base_info_to_mysql(dicts)
    
    dicts = excel_to_dict(excel_path, 'FWTOOLS-Status')
    base_info_to_mysql(dicts)
    
    dicts = excel_to_dict(excel_path, 'Factory-Status')
    factory_info_to_mysql(dicts)
    
    dicts = excel_to_dict(excel_path, 'Flow-Status')
    flow_info_to_mysql(dicts)
    
    filter_qti = 'assignee = bhuang AND component in (PROTOCOL, CONFIG, Configuration, CONFIGURATION) AND status not in (closed, done, integrated) AND project in (qti9x28, cougar, eel, eland, qti9x40, coronado, EMU) ORDER BY status ASC, summary DESC, key ASC'
    qti_to_mysql(filter_qti, QTI_List)